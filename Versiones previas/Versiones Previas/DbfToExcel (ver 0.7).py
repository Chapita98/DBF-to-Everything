import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import os
import dbfread
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment, PatternFill

# Limites de Excel = 1.048.576 filas
MAX_FILAS_EXCEL = 1000000

# Función común para aplicar formato
def aplicar_formato_excel(path, options):
    wb = load_workbook(path)
    ws = wb.active
    if options.get('center'):
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(horizontal='center', vertical='center')
    if options.get('freeze'):
        ws.freeze_panes = 'A2'
    if options.get('alternate'):
        for i, row in enumerate(ws.iter_rows(min_row=2), start=2):
            color = 'FFFFFF' if i % 2 == 0 else 'F0F0F0'
            for cell in row:
                cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
    if options.get('auto_size'):
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = max_len + 2
    wb.save(path)

# Preguntar cómo dividir un archivo muy grande
def preguntar_division():
    return messagebox.askyesno("Archivo muy grande", "El archivo supera el límite de filas de Excel (1.048.576).\n¿Deseas dividirlo en varias hojas dentro del mismo archivo?\n(Si eliges 'No', se dividirá en varios archivos.)")

# Guardar fragmentado
def guardar_fragmentado(df, ruta_base, aplicar_formato, opciones):
    partes = (len(df) + MAX_FILAS_EXCEL - 1) // MAX_FILAS_EXCEL
    usar_hojas = preguntar_division()

    if usar_hojas:
        wb = Workbook()
        for i in range(partes):
            parte = df.iloc[i * MAX_FILAS_EXCEL:(i + 1) * MAX_FILAS_EXCEL]
            ws = wb.create_sheet(title=f"Parte {i+1}") if i > 0 else wb.active
            ws.title = f"Parte {i+1}"
            ws.append(list(parte.columns))
            for fila in parte.itertuples(index=False):
                ws.append(fila)
        ruta_final = f"{ruta_base}.xlsx"
        wb.save(ruta_final)
        if aplicar_formato:
            aplicar_formato_excel(ruta_final, opciones)
    else:
        for i in range(partes):
            parte = df.iloc[i * MAX_FILAS_EXCEL:(i + 1) * MAX_FILAS_EXCEL]
            nombre_archivo = f"{ruta_base}_parte{i+1}.xlsx"
            parte.to_excel(nombre_archivo, index=False)
            if aplicar_formato:
                aplicar_formato_excel(nombre_archivo, opciones)

# Conversión básica
def convertir_basico():
    inp = basico_inp_var.get()
    out = basico_out_var.get()
    if not inp or not out:
        messagebox.showerror("Error", "Selecciona carpeta de entrada y salida.")
        return
    seleccion = [basico_tree.item(i)['values'][0] for i in basico_tree.selection()]
    if not seleccion:
        messagebox.showerror("Error", "Selecciona al menos un archivo.")
        return
    for nombre in seleccion:
        ruta_in = os.path.join(inp, nombre)
        ruta_out_base = os.path.splitext(os.path.join(out, nombre))[0]
        try:
            tabla = dbfread.DBF(ruta_in, encoding='latin1')
            df = pd.DataFrame(iter(tabla))
            if len(df) > MAX_FILAS_EXCEL:
                guardar_fragmentado(df, ruta_out_base, aplicar_formato=False, opciones={})
            else:
                df.to_excel(f"{ruta_out_base}.xlsx", index=False)
        except Exception as e:
            messagebox.showerror("Error", f"Falló {nombre}: {e}")
            return
    messagebox.showinfo("Éxito", "Archivos convertidos correctamente.")

# Conversión avanzada
def convertir_avanzado():
    inp = adv_inp_var.get()
    out = adv_out_var.get()
    tpl = adv_tpl_var.get()
    if not inp or not out:
        messagebox.showerror("Error", "Selecciona carpeta de entrada y salida.")
        return
    seleccion = [adv_tree.item(i)['values'][0] for i in adv_tree.selection()]
    if not seleccion:
        messagebox.showerror("Error", "Selecciona al menos un archivo.")
        return
    options = {
        'center': adv_center_var.get(),
        'freeze': adv_freeze_var.get(),
        'alternate': adv_alt_var.get(),
        'auto_size': adv_autosize_var.get()
    }
    for nombre in seleccion:
        ruta_in = os.path.join(inp, nombre)
        ruta_out_base = os.path.splitext(os.path.join(out, nombre))[0]
        try:
            tabla = dbfread.DBF(ruta_in, encoding='latin1')
            df = pd.DataFrame(iter(tabla))
            if tpl:
                wb = load_workbook(tpl)
                ws = wb.active
                ws.delete_rows(1, ws.max_row)
                for c, col in enumerate(df.columns, start=1):
                    ws.cell(row=1, column=c, value=col)
                for r, row in enumerate(df.itertuples(index=False), start=2):
                    for c, val in enumerate(row, start=1):
                        ws.cell(row=r, column=c, value=val)
                wb.save(f"{ruta_out_base}.xlsx")
                aplicar_formato_excel(f"{ruta_out_base}.xlsx", options)
            else:
                if len(df) > MAX_FILAS_EXCEL:
                    guardar_fragmentado(df, ruta_out_base, aplicar_formato=True, opciones=options)
                else:
                    df.to_excel(f"{ruta_out_base}.xlsx", index=False)
                    aplicar_formato_excel(f"{ruta_out_base}.xlsx", options)
        except Exception as e:
            messagebox.showerror("Error", f"Falló {nombre}: {e}")
            return
    messagebox.showinfo("Éxito", "Archivos convertidos con formato.")

# Conversión total
def convertir_total():
    inp = total_inp_var.get()
    out = total_out_var.get()
    if not inp or not out:
        messagebox.showerror("Error", "Selecciona carpeta de entrada y salida.")
        return
    files = [f for f in os.listdir(inp) if f.lower().endswith('.dbf')]
    options = {'center': True, 'freeze': True, 'alternate': True, 'auto_size': True} if total_fmt_var.get() else {}
    for nombre in files:
        ruta_in = os.path.join(inp, nombre)
        ruta_out_base = os.path.splitext(os.path.join(out, nombre))[0]
        try:
            tabla = dbfread.DBF(ruta_in, encoding='latin1')
            df = pd.DataFrame(iter(tabla))
            if len(df) > MAX_FILAS_EXCEL:
                guardar_fragmentado(df, ruta_out_base, aplicar_formato=bool(options), opciones=options)
            else:
                df.to_excel(f"{ruta_out_base}.xlsx", index=False)
                if options:
                    aplicar_formato_excel(f"{ruta_out_base}.xlsx", options)
        except Exception as e:
            print(f"Error {nombre}: {e}")
    messagebox.showinfo("Éxito", "Todos los archivos han sido convertidos.")
