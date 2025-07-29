import tkinter as tk
from tkinter import ttk, messagebox
import os
import dbfread
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, PatternFill
import tkfilebrowser

# Función para aplicar formato al Excel
def aplicar_formato_excel(ruta_salida):
    wb = load_workbook(ruta_salida)
    ws = wb.active
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.freeze_panes = 'A2'
    for i, row in enumerate(ws.iter_rows(min_row=2), start=2):
        color = 'FFFFFF' if i % 2 == 0 else 'F0F0F0'
        for cell in row:
            cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
    ws.row_dimensions[1].height = 25
    for col in ws.columns:
        max_length = max(len(str(cell.value or '')) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max_length + 2
    wb.save(ruta_salida)

# Convertidor básico
def convertir_basico():
    archivos = tkfilebrowser.askopenfilenames(filetypes=[("DBF files", "*.dbf")])
    if not archivos:
        return
    for archivo in archivos:
        try:
            tabla = dbfread.DBF(archivo, encoding='latin1')
            df = pd.DataFrame(iter(tabla))
            nombre_salida = os.path.splitext(archivo)[0] + '.xlsx'
            df.to_excel(nombre_salida, index=False)
            messagebox.showinfo("Éxito", f"Convertido: {nombre_salida}")
        except Exception as e:
            messagebox.showerror("Error", f"Error al procesar {archivo}: {e}")

# Convertidor con formato
def convertir_con_formato():
    archivos = tkfilebrowser.askopenfilenames(filetypes=[("DBF files", "*.dbf")])
    if not archivos:
        return
    for archivo in archivos:
        try:
            tabla = dbfread.DBF(archivo, encoding='latin1')
            df = pd.DataFrame(iter(tabla))
            nombre_salida = os.path.splitext(archivo)[0] + '.xlsx'
            df.to_excel(nombre_salida, index=False)
            aplicar_formato_excel(nombre_salida)
            messagebox.showinfo("Éxito", f"Convertido con formato: {nombre_salida}")
        except Exception as e:
            messagebox.showerror("Error", f"Error al procesar {archivo}: {e}")

# Conversor total
def convertir_total():
    directorio_entrada = entrada_total_var.get()
    directorio_salida = salida_total_var.get()
    if not directorio_entrada or not directorio_salida:
        messagebox.showerror("Error", "Por favor, selecciona ambos directorios.")
        return

    archivos_dbf = [f for f in os.listdir(directorio_entrada) if f.lower().endswith('.dbf')]
    cantidad = len(archivos_dbf)
    messagebox.showinfo("Información", f"Se encontraron {cantidad} archivos .dbf.")

    aplicar_formato = formato_total_var.get()

    for archivo in archivos_dbf:
        try:
            tabla = dbfread.DBF(os.path.join(directorio_entrada, archivo), encoding='latin1')
            df = pd.DataFrame(iter(tabla))
            nombre_salida = os.path.splitext(archivo)[0] + '.xlsx'
            ruta_salida = os.path.join(directorio_salida, nombre_salida)
            df.to_excel(ruta_salida, index=False)
            if aplicar_formato:
                aplicar_formato_excel(ruta_salida)
            print(f"Convertido: {ruta_salida}")
        except Exception as e:
            print(f"Error al procesar {archivo}: {e}")

    messagebox.showinfo("Éxito", "Todos los archivos han sido convertidos.")

# Crear ventana principal
ventana = tk.Tk()
ventana.title("Convertir DBF a Excel")
ventana.geometry("600x400")

# Crear contenedor de pestañas
notebook = ttk.Notebook(ventana)
notebook.pack(fill='both', expand=True)

# Pestaña 1: Convertidor Básico
pestana_basico = ttk.Frame(notebook)
notebook.add(pestana_basico, text='Convertidor Básico')
tk.Label(pestana_basico, text="Selecciona archivos .dbf para convertir:").pack(pady=10)
tk.Button(pestana_basico, text="Seleccionar y Convertir", command=convertir_basico).pack(pady=10)

# Pestaña 2: Convertidor con Formato
pestana_formato = ttk.Frame(notebook)
notebook.add(pestana_formato, text='Convertidor con Formato')
tk.Label(pestana_formato, text="Selecciona archivos .dbf para convertir con formato:").pack(pady=10)
tk.Button(pestana_formato, text="Seleccionar y Convertir", command=convertir_con_formato).pack(pady=10)

# Pestaña 3: Conversor Total
pestana_total = ttk.Frame(notebook)
notebook.add(pestana_total, text='Conversor Total')

tk.Label(pestana_total, text="Directorio de entrada:").pack(pady=5)
entrada_total_var = tk.StringVar()
tk.Entry(pestana_total, textvariable=entrada_total_var, width=50).pack(pady=5)
tk.Button(pestana_total, text="Seleccionar", command=lambda: entrada_total_var.set(tkfilebrowser.askopendirname())).pack(pady=5)

tk.Label(pestana_total, text="Directorio de salida:").pack(pady=5)
salida_total_var = tk.StringVar()
tk.Entry(pestana_total, textvariable=salida_total_var, width=50).pack(pady=5)
tk.Button(pestana_total, text="Seleccionar", command=lambda: salida_total_var.set(tkfilebrowser.askopendirname())).pack(pady=5)

formato_total_var = tk.BooleanVar()
tk.Checkbutton(pestana_total, text="Aplicar formato a los archivos", variable=formato_total_var).pack(pady=10)

tk.Button(pestana_total, text="Convertir Todos", command=convertir_total).pack(pady=10)

# Iniciar la ventana
ventana.mainloop()