import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import os
import dbfread
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, PatternFill

# Función común para aplicar formato
def aplicar_formato_excel(path, options):
    wb = load_workbook(path)
    ws = wb.active
    # Centrar texto
    if options.get('center'):
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(horizontal='center', vertical='center')
    # Inmovilizar primera fila
    if options.get('freeze'):
        ws.freeze_panes = 'A2'
    # Colores alternados
    if options.get('alternate'):
        for i, row in enumerate(ws.iter_rows(min_row=2), start=2):
            color = 'FFFFFF' if i % 2 == 0 else 'F0F0F0'
            for cell in row:
                cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
    # Ajustar ancho columnas
    if options.get('auto_size'):
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = max_len + 2
    wb.save(path)

# Conversión básica
def convertir_basico():
    inp = basico_inp_var.get()
    out = basico_out_var.get()
    if not inp or not out:
        messagebox.showerror("Error", "Selecciona carpeta de entrada y salida.")
        return
    seleccion = [basico_tree.item(i)['values'][0] for i in basico_tree.selection()]
    if not seleccion:
        messagebox.showerror("Error", "Selecciona al menos un archivo de la lista.")
        return
    for nombre in seleccion:
        ruta_in = os.path.join(inp, nombre)
        ruta_out = os.path.join(out, os.path.splitext(nombre)[0] + '.xlsx')
        try:
            tabla = dbfread.DBF(ruta_in, encoding='latin1')
            df = pd.DataFrame(iter(tabla))
            df.to_excel(ruta_out, index=False)
        except Exception as e:
            messagebox.showerror("Error", f"Falló {nombre}: {e}")
            return
    messagebox.showinfo("Éxito", "Archivos convertidos correctamente.")

# Recargar lista básica
def cargar_lista_basico():
    basico_tree.delete(*basico_tree.get_children())
    inp = basico_inp_var.get()
    if not inp:
        return
    for f in os.listdir(inp):
        if f.lower().endswith('.dbf'):
            basico_tree.insert('', 'end', values=(f,))

# Conversión avanzada
def convertir_avanzado():
    inp = adv_inp_var.get()
    out = adv_out_var.get()
    tpl = adv_tpl_var.get()
    if not inp or not out:
        messagebox.showerror("Error", "Selecciona carpeta de entrada y salida.")
        return
    seleccion = [adv_tree.item(i)['values'][0] for i in adv_tree.selection()]
    if not seleccion:
        messagebox.showerror("Error", "Selecciona al menos un archivo.")
        return
    options = {
        'center': adv_center_var.get(),
        'freeze': adv_freeze_var.get(),
        'alternate': adv_alt_var.get(),
        'auto_size': adv_autosize_var.get()
    }
    for nombre in seleccion:
        ruta_in = os.path.join(inp, nombre)
        nombre_out = os.path.splitext(nombre)[0] + '.xlsx'
        ruta_out = os.path.join(out, nombre_out)
        try:
            tabla = dbfread.DBF(ruta_in, encoding='latin1')
            df = pd.DataFrame(iter(tabla))
            if tpl:
                wb = load_workbook(tpl)
                ws = wb.active
                # limpiar contenido
                ws.delete_rows(1, ws.max_row)
                # escribir encabezados
                for c, col in enumerate(df.columns, start=1):
                    ws.cell(row=1, column=c, value=col)
                # escribir datos
                for r, row in enumerate(df.itertuples(index=False), start=2):
                    for c, val in enumerate(row, start=1):
                        ws.cell(row=r, column=c, value=val)
                wb.save(ruta_out)
            else:
                df.to_excel(ruta_out, index=False)
            aplicar_formato_excel(ruta_out, options)
        except Exception as e:
            messagebox.showerror("Error", f"Falló {nombre}: {e}")
            return
    messagebox.showinfo("Éxito", "Archivos convertidos con formato.")

# Recargar lista avanzada
def cargar_lista_adv():
    adv_tree.delete(*adv_tree.get_children())
    inp = adv_inp_var.get()
    if not inp:
        return
    for f in os.listdir(inp):
        if f.lower().endswith('.dbf'):
            adv_tree.insert('', 'end', values=(f,))

# Conversor total
def convertir_total():
    inp = total_inp_var.get()
    out = total_out_var.get()
    if not inp or not out:
        messagebox.showerror("Error", "Selecciona carpeta de entrada y salida.")
        return
    files = [f for f in os.listdir(inp) if f.lower().endswith('.dbf')]
    cnt = len(files)
    messagebox.showinfo("Info", f"Se encontraron {cnt} archivos .dbf.")
    options = {'center': True, 'freeze': True, 'alternate': True, 'auto_size': True} if total_fmt_var.get() else {}
    for nombre in files:
        ruta_in = os.path.join(inp, nombre)
        ruta_out = os.path.join(out, os.path.splitext(nombre)[0] + '.xlsx')
        try:
            tabla = dbfread.DBF(ruta_in, encoding='latin1')
            df = pd.DataFrame(iter(tabla))
            df.to_excel(ruta_out, index=False)
            if options:
                aplicar_formato_excel(ruta_out, options)
        except Exception as e:
            print(f"Error {nombre}: {e}")
    messagebox.showinfo("Éxito", "Todos los archivos han sido convertidos.")

# --------------------- UI ---------------------
root = tk.Tk()
root.title("Convertidor DBF a Excel")
root.geometry("700x500")
notebook = ttk.Notebook(root)
notebook.pack(fill='both', expand=True)

# Pestaña Básico
frame1 = ttk.Frame(notebook)
notebook.add(frame1, text='Convertidor Básico')
basico_inp_var = tk.StringVar()
basico_out_var = tk.StringVar()
# Directorios
ttk.Label(frame1, text="Carpeta entrada:").grid(row=0, column=0, pady=5, sticky='w')
ttk.Entry(frame1, textvariable=basico_inp_var, width=50).grid(row=0, column=1)
ttk.Button(frame1, text="Examinar", command=lambda: basico_inp_var.set(filedialog.askdirectory())).grid(row=0, column=2)

ttk.Label(frame1, text="Carpeta salida:").grid(row=1, column=0, pady=5, sticky='w')
ttk.Entry(frame1, textvariable=basico_out_var, width=50).grid(row=1, column=1)
ttk.Button(frame1, text="Examinar", command=lambda: basico_out_var.set(filedialog.askdirectory())).grid(row=1, column=2)

# Lista archivos
ttk.Button(frame1, text="Cargar lista", command=cargar_lista_basico).grid(row=2, column=0, pady=10)
basico_tree = ttk.Treeview(frame1, columns=('archivo',), show='headings', height=10)
basico_tree.heading('archivo', text='Archivo .dbf')
basico_tree.grid(row=3, column=0, columnspan=3, padx=10)

ttk.Button(frame1, text="Convertir Seleccionados", command=convertir_basico).grid(row=4, column=1, pady=10)

# Pestaña Avanzado
frame2 = ttk.Frame(notebook)
notebook.add(frame2, text='Convertidor Avanzado')
adv_inp_var = tk.StringVar()
adv_out_var = tk.StringVar()
adv_tpl_var = tk.StringVar()
adv_center_var = tk.BooleanVar()
adv_freeze_var = tk.BooleanVar()
adv_alt_var = tk.BooleanVar()
adv_autosize_var = tk.BooleanVar()
# Directorios y plantilla
ttk.Label(frame2, text="Carpeta entrada:").grid(row=0, column=0, pady=5, sticky='w')
ttk.Entry(frame2, textvariable=adv_inp_var, width=50).grid(row=0, column=1)
ttk.Button(frame2, text="Examinar", command=lambda: adv_inp_var.set(filedialog.askdirectory())).grid(row=0, column=2)

ttk.Label(frame2, text="Carpeta salida:").grid(row=1, column=0, pady=5, sticky='w')
ttk.Entry(frame2, textvariable=adv_out_var, width=50).grid(row=1, column=1)
ttk.Button(frame2, text="Examinar", command=lambda: adv_out_var.set(filedialog.askdirectory())).grid(row=1, column=2)

ttk.Label(frame2, text="Plantilla Excel (opcional):").grid(row=2, column=0, pady=5, sticky='w')
ttk.Entry(frame2, textvariable=adv_tpl_var, width=50).grid(row=2, column=1)
ttk.Button(frame2, text="Seleccionar", command=lambda: adv_tpl_var.set(filedialog.askopenfilename(filetypes=[('Excel', '*.xlsx')]))).grid(row=2, column=2)

# Opciones de formato
ttk.Checkbutton(frame2, text="Centrar texto", variable=adv_center_var).grid(row=3, column=0)
ttk.Checkbutton(frame2, text="Inmovilizar primera fila", variable=adv_freeze_var).grid(row=3, column=1)
ttk.Checkbutton(frame2, text="Colores alternados", variable=adv_alt_var).grid(row=3, column=2)
ttk.Checkbutton(frame2, text="Auto-ajustar columnas", variable=adv_autosize_var).grid(row=4, column=0)

# Lista archivos avanzados
ttk.Button(frame2, text="Cargar lista", command=cargar_lista_adv).grid(row=5, column=0, pady=10)
adv_tree = ttk.Treeview(frame2, columns=('archivo',), show='headings', height=8)
adv_tree.heading('archivo', text='Archivo .dbf')
adv_tree.grid(row=6, column=0, columnspan=3, padx=10)

ttk.Button(frame2, text="Convertir Seleccionados", command=convertir_avanzado).grid(row=7, column=1, pady=10)

# Pestaña Total
frame3 = ttk.Frame(notebook)
notebook.add(frame3, text='Convertidor Total')
total_inp_var = tk.StringVar()
total_out_var = tk.StringVar()
total_fmt_var = tk.BooleanVar()
# Directorios
ttk.Label(frame3, text="Carpeta entrada:").grid(row=0, column=0, pady=5, sticky='w')
ttk.Entry(frame3, textvariable=total_inp_var, width=50).grid(row=0, column=1)
ttk.Button(frame3, text="Examinar", command=lambda: [total_inp_var.set(filedialog.askdirectory()), actualizar_contador()]).grid(row=0, column=2)

ttk.Label(frame3, text="Carpeta salida:").grid(row=1, column=0, pady=5, sticky='w')
ttk.Entry(frame3, textvariable=total_out_var, width=50).grid(row=1, column=1)
ttk.Button(frame3, text="Examinar", command=lambda: total_out_var.set(filedialog.askdirectory())).grid(row=1, column=2)

# Contador archivos
total_count_lbl = ttk.Label(frame3, text="Archivos .dbf encontrados: 0")
total_count_lbl.grid(row=2, column=0, columnspan=2, pady=5, sticky='w')

def actualizar_contador():
    inp = total_inp_var.get()
    cnt = len([f for f in os.listdir(inp) if f.lower().endswith('.dbf')]) if inp else 0
    total_count_lbl.config(text=f"Archivos .dbf encontrados: {cnt}")

# Opción de formato masivo
ttk.Checkbutton(frame3, text="Aplicar formato a todos", variable=total_fmt_var).grid(row=3, column=0, columnspan=2, pady=5)
 
# Botón convertir total
ttk.Button(frame3, text="Convertir Todos", command=convertir_total).grid(row=4, column=1, pady=10)

root.mainloop()
