import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import os
import dbfread
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment, PatternFill
from datetime import datetime

# Limites de Excel = 1.048.576 filas
MAX_FILAS_EXCEL = 1000000



# Ruta del log
LOG_PATH = os.path.join("logs", "registro_conversiones.txt")

# Función para registrar en el log
def registrar_log(nombre_archivo, estado, detalles=""):
    os.makedirs("logs", exist_ok=True)
    with open(LOG_PATH, "a", encoding="utf-8") as f:
        ahora = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        f.write(f"[{ahora}] {nombre_archivo} - {estado} - {detalles}\n")

# Función común para aplicar formato
def aplicar_formato_excel(path, options):
    wb = load_workbook(path)
    ws = wb.active
    if options.get('center'):
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(horizontal='center', vertical='center')
    if options.get('freeze'):
        ws.freeze_panes = 'A2'
    if options.get('alternate'):
        for i, row in enumerate(ws.iter_rows(min_row=2), start=2):
            color = 'FFFFFF' if i % 2 == 0 else 'F0F0F0'
            for cell in row:
                cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
    if options.get('auto_size'):
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = max_len + 2
    wb.save(path)

# Preguntar cómo dividir un archivo muy grande
def preguntar_division():
    return messagebox.askyesno(
        "Archivo muy grande",
        "El archivo supera el límite de filas de Excel (1.048.576).\n"
        "¿Deseas dividirlo en varias hojas dentro del mismo archivo?\n"
        "(Si eliges 'No', se dividirá en varios archivos.)"
    )

# Preguntar si mantener nombres originales
def preguntar_nombres():
    return messagebox.askyesno(
        "Nombres personalizados",
        "¿Deseas mantener los nombres exactos de los archivos DBF para los archivos Excel?\n"
        "(Si eliges 'No', los nombres que comiencen con 'x' tendrán esa letra eliminada.)"
    )

# Obtener nombre base sin la letra x si se desea
def obtener_nombre_base(nombre):
    if not mantener_nombres_var.get():
        if eliminar_x_var.get() and nombre.lower().startswith('x'):
            nombre = nombre[1:]
    return os.path.splitext(nombre)[0]

# Guardar archivo fragmentado
def guardar_fragmentado(df, ruta_base, aplicar_formato, opciones):
    partes = (len(df) + MAX_FILAS_EXCEL - 1) // MAX_FILAS_EXCEL
    usar_hojas = preguntar_division()

    if usar_hojas:
        wb = Workbook()
        for i in range(partes):
            parte = df.iloc[i * MAX_FILAS_EXCEL:(i + 1) * MAX_FILAS_EXCEL]
            ws = wb.create_sheet(title=f"Parte {i+1}") if i > 0 else wb.active
            ws.title = f"Parte {i+1}"
            ws.append(list(parte.columns))
            for fila in parte.itertuples(index=False):
                ws.append(fila)
        ruta_final = f"{ruta_base}.xlsx"
        wb.save(ruta_final)
        if aplicar_formato:
            aplicar_formato_excel(ruta_final, opciones)
        registrar_log(os.path.basename(ruta_final), "OK", "Fragmentado en hojas")
    else:
        for i in range(partes):
            parte = df.iloc[i * MAX_FILAS_EXCEL:(i + 1) * MAX_FILAS_EXCEL]
            nombre_archivo = f"{ruta_base}_parte{i+1}.xlsx"
            parte.to_excel(nombre_archivo, index=False)
            if aplicar_formato:
                aplicar_formato_excel(nombre_archivo, opciones)
            registrar_log(os.path.basename(nombre_archivo), "OK", f"Parte {i+1} de {partes}")


# Conversión básica
def convertir_basico():
    mantener = preguntar_nombres()
    mantener_nombres_var.set(mantener)

    inp = basico_inp_var.get()
    out = basico_out_var.get()
    seleccion = [basico_tree.item(i)['values'][0] for i in basico_tree.selection()]
    if not inp or not out or not seleccion:
        messagebox.showerror("Error", "Completa todos los campos y selecciona archivos.")
        return

    progreso['maximum'] = len(seleccion)
    progreso['value'] = 0

    for nombre in seleccion:
        ruta_in = os.path.join(inp, nombre)
        nombre_base = obtener_nombre_base(nombre)
        ruta_out_base = os.path.join(out, nombre_base)

        try:
            df = pd.DataFrame(iter(dbfread.DBF(ruta_in, encoding='latin1')))
            if len(df) > MAX_FILAS_EXCEL:
                guardar_fragmentado(df, ruta_out_base, aplicar_formato=False, opciones={})
            else:
                df.to_excel(f"{ruta_out_base}.xlsx", index=False)
                registrar_log(nombre, "OK", "Básico")
        except Exception as e:
            registrar_log(nombre, "ERROR", str(e))
            messagebox.showerror("Error", f"Falló {nombre}: {e}")
        progreso['value'] += 1
        root.update_idletasks()

    messagebox.showinfo("Éxito", "Archivos convertidos.")

# Conversión avanzada
def convertir_avanzado():
    mantener = preguntar_nombres()
    mantener_nombres_var.set(mantener)

    inp = adv_inp_var.get()
    out = adv_out_var.get()
    tpl = adv_tpl_var.get()
    seleccion = [adv_tree.item(i)['values'][0] for i in adv_tree.selection()]
    if not inp or not out or not seleccion:
        messagebox.showerror("Error", "Completa todos los campos y selecciona archivos.")
        return

    options = {
        'center': adv_center_var.get(),
        'freeze': adv_freeze_var.get(),
        'alternate': adv_alt_var.get(),
        'auto_size': adv_autosize_var.get()
    }

    progreso['maximum'] = len(seleccion)
    progreso['value'] = 0

    for nombre in seleccion:
        ruta_in = os.path.join(inp, nombre)
        nombre_base = obtener_nombre_base(nombre)
        ruta_out_base = os.path.join(out, nombre_base)

        try:
            df = pd.DataFrame(iter(dbfread.DBF(ruta_in, encoding='latin1')))
            if tpl:
                wb = load_workbook(tpl)
                ws = wb.active
                ws.delete_rows(1, ws.max_row)
                for c, col in enumerate(df.columns, start=1):
                    ws.cell(row=1, column=c, value=col)
                for r, row in enumerate(df.itertuples(index=False), start=2):
                    for c, val in enumerate(row, start=1):
                        ws.cell(row=r, column=c, value=val)
                ruta_final = f"{ruta_out_base}.xlsx"
                wb.save(ruta_final)
                aplicar_formato_excel(ruta_final, options)
                registrar_log(nombre, "OK", "Con plantilla")
            else:
                if len(df) > MAX_FILAS_EXCEL:
                    guardar_fragmentado(df, ruta_out_base, aplicar_formato=True, opciones=options)
                else:
                    df.to_excel(f"{ruta_out_base}.xlsx", index=False)
                    aplicar_formato_excel(f"{ruta_out_base}.xlsx", options)
                    registrar_log(nombre, "OK", "Avanzado")

        except Exception as e:
            registrar_log(nombre, "ERROR", str(e))
            messagebox.showerror("Error", f"Falló {nombre}: {e}")
        progreso['value'] += 1
        root.update_idletasks()

    messagebox.showinfo("Éxito", "Archivos convertidos con formato.")


# Conversión total
def convertir_total():
    mantener = preguntar_nombres()
    mantener_nombres_var.set(mantener)

    inp = total_inp_var.get()
    out = total_out_var.get()
    if not inp or not out:
        messagebox.showerror("Error", "Selecciona carpeta de entrada y salida.")
        return

    files = [f for f in os.listdir(inp) if f.lower().endswith('.dbf')]
    options = {
        'center': True,
        'freeze': True,
        'alternate': True,
        'auto_size': True
    } if total_fmt_var.get() else {}

    progreso['maximum'] = len(files)
    progreso['value'] = 0

    for nombre in files:
        ruta_in = os.path.join(inp, nombre)
        nombre_base = obtener_nombre_base(nombre)
        ruta_out_base = os.path.join(out, nombre_base)

        try:
            df = pd.DataFrame(iter(dbfread.DBF(ruta_in, encoding='latin1')))
            if len(df) > MAX_FILAS_EXCEL:
                guardar_fragmentado(df, ruta_out_base, aplicar_formato=bool(options), opciones=options)
            else:
                df.to_excel(f"{ruta_out_base}.xlsx", index=False)
                if options:
                    aplicar_formato_excel(f"{ruta_out_base}.xlsx", options)
                registrar_log(nombre, "OK", "Total")
        except Exception as e:
            registrar_log(nombre, "ERROR", str(e))
        progreso['value'] += 1
        root.update_idletasks()

    messagebox.showinfo("Éxito", "Todos los archivos han sido convertidos.")



# ------------------- Interfaz gráfica -------------------

# Interfaz principal
root = tk.Tk()
root.title("Convertidor DBF a Excel")
root.geometry("760x560")

# Crear variables Tkinter después del root
mantener_nombres_var = tk.BooleanVar(value=True)
eliminar_x_var = tk.BooleanVar(value=False)



notebook = ttk.Notebook(root)
notebook.pack(fill='both', expand=True)

# Barra de progreso al fondo
progreso = ttk.Progressbar(root, length=740, mode='determinate')
progreso.pack(pady=5)

frame_botones = ttk.Frame(root)
frame_botones.pack()
ttk.Checkbutton(frame_botones, text="Eliminar letra 'x' del inicio del nombre", variable=eliminar_x_var).pack(side='left', padx=10, pady=5)

# Convertidor Básico
frame1 = ttk.Frame(notebook)
notebook.add(frame1, text='Convertidor Básico')

basico_inp_var = tk.StringVar()
basico_out_var = tk.StringVar()

ttk.Label(frame1, text="Carpeta de entrada:").grid(row=0, column=0, sticky='w', padx=5, pady=5)
ttk.Entry(frame1, textvariable=basico_inp_var, width=50).grid(row=0, column=1)
ttk.Button(frame1, text="Examinar", command=lambda: basico_inp_var.set(filedialog.askdirectory())).grid(row=0, column=2)

ttk.Label(frame1, text="Carpeta de salida:").grid(row=1, column=0, sticky='w', padx=5, pady=5)
ttk.Entry(frame1, textvariable=basico_out_var, width=50).grid(row=1, column=1)
ttk.Button(frame1, text="Examinar", command=lambda: basico_out_var.set(filedialog.askdirectory())).grid(row=1, column=2)

ttk.Button(frame1, text="Cargar lista", command=lambda: cargar_lista_archivos(basico_tree, basico_inp_var)).grid(row=2, column=0, pady=10)

basico_tree = ttk.Treeview(frame1, columns=('archivo',), show='headings', height=10)
basico_tree.heading('archivo', text='Archivo .dbf')
basico_tree.grid(row=3, column=0, columnspan=3, padx=10)

ttk.Button(frame1, text="Convertir seleccionados", command=convertir_basico).grid(row=4, column=1, pady=10)


# Función para cargar lista de archivos en Treeview
def cargar_lista_archivos(treeview, carpeta_var):
    treeview.delete(*treeview.get_children())
    carpeta = carpeta_var.get()
    if not carpeta:
        return
    for f in os.listdir(carpeta):
        if f.lower().endswith('.dbf'):
            treeview.insert('', 'end', values=(f,))

# Convertidor Avanzado
frame2 = ttk.Frame(notebook)
notebook.add(frame2, text='Convertidor Avanzado')

adv_inp_var = tk.StringVar()
adv_out_var = tk.StringVar()
adv_tpl_var = tk.StringVar()
adv_center_var = tk.BooleanVar()
adv_freeze_var = tk.BooleanVar()
adv_alt_var = tk.BooleanVar()
adv_autosize_var = tk.BooleanVar()

ttk.Label(frame2, text="Carpeta de entrada:").grid(row=0, column=0, sticky='w', padx=5, pady=5)
ttk.Entry(frame2, textvariable=adv_inp_var, width=50).grid(row=0, column=1)
ttk.Button(frame2, text="Examinar", command=lambda: adv_inp_var.set(filedialog.askdirectory())).grid(row=0, column=2)

ttk.Label(frame2, text="Carpeta de salida:").grid(row=1, column=0, sticky='w', padx=5, pady=5)
ttk.Entry(frame2, textvariable=adv_out_var, width=50).grid(row=1, column=1)
ttk.Button(frame2, text="Examinar", command=lambda: adv_out_var.set(filedialog.askdirectory())).grid(row=1, column=2)

ttk.Label(frame2, text="Plantilla Excel (opcional):").grid(row=2, column=0, sticky='w', padx=5, pady=5)
ttk.Entry(frame2, textvariable=adv_tpl_var, width=50).grid(row=2, column=1)
ttk.Button(frame2, text="Seleccionar", command=lambda: adv_tpl_var.set(filedialog.askopenfilename(filetypes=[('Excel', '*.xlsx')]))).grid(row=2, column=2)

ttk.Checkbutton(frame2, text="Centrar texto", variable=adv_center_var).grid(row=3, column=0)
ttk.Checkbutton(frame2, text="Inmovilizar primera fila", variable=adv_freeze_var).grid(row=3, column=1)
ttk.Checkbutton(frame2, text="Colores alternados", variable=adv_alt_var).grid(row=3, column=2)
ttk.Checkbutton(frame2, text="Auto-ajustar columnas", variable=adv_autosize_var).grid(row=4, column=0)

ttk.Button(frame2, text="Cargar lista", command=lambda: cargar_lista_archivos(adv_tree, adv_inp_var)).grid(row=5, column=0, pady=10)

adv_tree = ttk.Treeview(frame2, columns=('archivo',), show='headings', height=8)
adv_tree.heading('archivo', text='Archivo .dbf')
adv_tree.grid(row=6, column=0, columnspan=3, padx=10)

ttk.Button(frame2, text="Convertir seleccionados", command=convertir_avanzado).grid(row=7, column=1, pady=10)


# Convertidor Total
frame3 = ttk.Frame(notebook)
notebook.add(frame3, text='Convertidor Total')

total_inp_var = tk.StringVar()
total_out_var = tk.StringVar()
total_fmt_var = tk.BooleanVar()

ttk.Label(frame3, text="Carpeta de entrada:").grid(row=0, column=0, sticky='w', padx=5, pady=5)
ttk.Entry(frame3, textvariable=total_inp_var, width=50).grid(row=0, column=1)
ttk.Button(frame3, text="Examinar", command=lambda: total_inp_var.set(filedialog.askdirectory())).grid(row=0, column=2)

ttk.Label(frame3, text="Carpeta de salida:").grid(row=1, column=0, sticky='w', padx=5, pady=5)
ttk.Entry(frame3, textvariable=total_out_var, width=50).grid(row=1, column=1)
ttk.Button(frame3, text="Examinar", command=lambda: total_out_var.set(filedialog.askdirectory())).grid(row=1, column=2)

ttk.Checkbutton(frame3, text="Aplicar formato a todos", variable=total_fmt_var).grid(row=2, column=0, columnspan=2, pady=10)

ttk.Button(frame3, text="Convertir Todos", command=convertir_total).grid(row=3, column=1, pady=10)


# Agregar todas las pestañas al cuaderno
notebook.add(frame1, text="Convertidor Básico")
notebook.add(frame2, text="Convertidor Avanzado")
notebook.add(frame3, text="Convertidor Total")

root.mainloop()
