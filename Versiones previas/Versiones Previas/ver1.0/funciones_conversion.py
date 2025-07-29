import os
import dbfread
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

def aplicar_formato_excel(archivo_salida, centrar_texto=False, inmovilizar_fila=False,
                          colores_alternados=False, autoajustar_columnas=False,
                          negrita=False, tamano_encabezado=12, tamano_datos=11):
    wb = openpyxl.load_workbook(archivo_salida)
    ws = wb.active

    fuente_encabezado = Font(bold=negrita, size=tamano_encabezado)
    fuente_datos = Font(bold=negrita, size=tamano_datos)
    alineacion_centrada = Alignment(horizontal='center', vertical='center')

    for col_idx, col in enumerate(ws.iter_cols(min_row=1, max_row=1), start=1):
        for cell in col:
            cell.font = fuente_encabezado
            if centrar_texto:
                cell.alignment = alineacion_centrada

    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        for cell in row:
            cell.font = fuente_datos
            if centrar_texto:
                cell.alignment = alineacion_centrada
            if colores_alternados and row_idx % 2 == 0:
                cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")

    if inmovilizar_fila:
        ws.freeze_panes = "A2"

    if autoajustar_columnas:
        for col in ws.columns:
            max_length = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = max_length + 2

    wb.save(archivo_salida)

def convertir_basico(archivo_dbf, carpeta_salida):
    nombre_archivo = os.path.splitext(os.path.basename(archivo_dbf))[0]
    archivo_salida = os.path.join(carpeta_salida, f"{nombre_archivo}.xlsx")
    tabla = dbfread.DBF(archivo_dbf, load=True)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(tabla.field_names)
    for registro in tabla:
        ws.append(list(registro.values()))

    wb.save(archivo_salida)
    return archivo_salida

def convertir_avanzado(archivo_dbf, carpeta_salida, opciones):
    nombre_archivo = os.path.splitext(os.path.basename(archivo_dbf))[0]
    archivo_salida = os.path.join(carpeta_salida, f"{nombre_archivo}.xlsx")
    tabla = dbfread.DBF(archivo_dbf, load=True)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(tabla.field_names)
    for registro in tabla:
        ws.append(list(registro.values()))

    wb.save(archivo_salida)

    aplicar_formato_excel(
        archivo_salida,
        centrar_texto=opciones["centrar_texto"],
        inmovilizar_fila=opciones["inmovilizar_fila"],
        colores_alternados=opciones["colores_alternados"],
        autoajustar_columnas=opciones["autoajustar_columnas"],
        negrita=opciones["negrita"],
        tamano_encabezado=opciones["tamano_encabezado"],
        tamano_datos=opciones["tamano_datos"]
    )
    return archivo_salida

def convertir_total(lista_archivos, carpeta_salida, opciones):
    resultados = []
    for archivo in lista_archivos:
        resultado = convertir_avanzado(archivo, carpeta_salida, opciones)
        resultados.append(resultado)
    return resultados
