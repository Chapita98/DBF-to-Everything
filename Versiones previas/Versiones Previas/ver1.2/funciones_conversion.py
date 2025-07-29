import os
import pandas as pd
from dbfread import DBF
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment, Font, PatternFill

from utilidades import obtener_nombre_archivo_sin_extension

def convertir_dbf_a_dataframe(ruta):
    dbf = DBF(ruta, load=True)
    return pd.DataFrame(iter(dbf))

def guardar_dataframe_como_excel(df, ruta_salida):
    wb = Workbook()
    ws = wb.active

    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)

    wb.save(ruta_salida)

def aplicar_personalizacion_excel(ruta_excel, opciones):
    from openpyxl import load_workbook

    wb = load_workbook(ruta_excel)
    ws = wb.active

    max_col = ws.max_column
    max_row = ws.max_row

    fuente_header = Font(bold=True, size=opciones.get("tamano_encabezado", 12))
    fuente_datos = Font(size=opciones.get("tamano_datos", 10))

    for col in range(1, max_col + 1):
        for row in range(1, max_row + 1):
            celda = ws.cell(row=row, column=col)

            # Estilo para encabezado
            if row == 1:
                celda.font = fuente_header
                if opciones.get("centrar_texto", True):
                    celda.alignment = Alignment(horizontal="center")
                if opciones.get("color_alternado", False):
                    celda.fill = PatternFill(start_color="D9D9D9", fill_type="solid")
            else:
                celda.font = fuente_datos
                if opciones.get("centrar_texto", True):
                    celda.alignment = Alignment(horizontal="center")

    if opciones.get("inmovilizar_fila_1", True):
        ws.freeze_panes = "A2"

    if opciones.get("autocolumnas", True):
        for col in ws.columns:
            max_length = 0
            columna = col[0].column_letter
            for celda in col:
                try:
                    if celda.value:
                        max_length = max(max_length, len(str(celda.value)))
                except:
                    pass
            ajuste = min((max_length + 2), 50)
            ws.column_dimensions[columna].width = ajuste

    wb.save(ruta_excel)

def convertir_basico(ruta_entrada, ruta_salida):
    df = convertir_dbf_a_dataframe(ruta_entrada)
    guardar_dataframe_como_excel(df, ruta_salida)

def convertir_avanzado(ruta_entrada, ruta_salida, opciones):
    df = convertir_dbf_a_dataframe(ruta_entrada)
    guardar_dataframe_como_excel(df, ruta_salida)
    aplicar_personalizacion_excel(ruta_salida, opciones)

def convertir_total(lista_rutas, ruta_salida_carpeta, opciones):
    for ruta_entrada in lista_rutas:
        nombre_archivo = obtener_nombre_archivo_sin_extension(ruta_entrada)
        ruta_salida = os.path.join(ruta_salida_carpeta, f"{nombre_archivo}.xlsx")
        convertir_avanzado(ruta_entrada, ruta_salida, opciones)
