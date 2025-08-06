import os
import pandas as pd
from dbfread import DBF
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl import load_workbook

from utilidades import obtener_nombre_archivo_sin_extension
from tkinter import messagebox

MAX_FILAS_EXCEL = 1_000_000

def convertir_dbf_a_dataframe(ruta):
    dbf = DBF(ruta, load=True)
    return pd.DataFrame(iter(dbf))

def guardar_dataframe_como_excel(df, ruta_salida, fragmento_hojas=False, fragmento_archivos=False, progreso_callback=None):
    if fragmento_hojas:
        wb = Workbook()
        df_chunks = [df[i:i+MAX_FILAS_EXCEL] for i in range(0, len(df), MAX_FILAS_EXCEL)]
        for idx, chunk in enumerate(df_chunks):
            ws = wb.create_sheet(title=f"Hoja{idx+1}")
            total_rows = len(chunk)
            for i, r in enumerate(dataframe_to_rows(chunk, index=False, header=True)):
                ws.append(r)
                if progreso_callback and i % 500 == 0:
                    progreso_callback(i+1, total_rows)
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
        wb.save(ruta_salida)
    elif fragmento_archivos:
        base, ext = os.path.splitext(ruta_salida)
        df_chunks = [df[i:i+MAX_FILAS_EXCEL] for i in range(0, len(df), MAX_FILAS_EXCEL)]
        for idx, chunk in enumerate(df_chunks):
            ruta_nueva = f"{base}_parte{idx+1}{ext}"
            wb = Workbook()
            ws = wb.active
            total_rows = len(chunk)
            for i, r in enumerate(dataframe_to_rows(chunk, index=False, header=True)):
                ws.append(r)
                if progreso_callback and i % 500 == 0:
                    progreso_callback(i+1, total_rows)
            wb.save(ruta_nueva)
    else:
        wb = Workbook()
        ws = wb.active
        total_rows = len(df)
        for i, r in enumerate(dataframe_to_rows(df, index=False, header=True)):
            ws.append(r)
            if progreso_callback and i % 500 == 0:
                progreso_callback(i+1, total_rows)
        wb.save(ruta_salida)

def aplicar_personalizacion_excel(ruta_excel, opciones):
    wb = load_workbook(ruta_excel)
    ws = wb.active

    max_col = ws.max_column
    max_row = ws.max_row

    fuente_header = Font(bold=True, size=opciones.get("tamano_encabezado", 12))
    fuente_datos = Font(size=opciones.get("tamano_datos", 10))

    for col in range(1, max_col + 1):
        for row in range(1, max_row + 1):
            celda = ws.cell(row=row, column=col)

            if row == 1:
                celda.font = fuente_header
                if opciones.get("centrar_texto", True):
                    celda.alignment = Alignment(horizontal="center")
                if opciones.get("color_alternado", False):
                    celda.fill = PatternFill(start_color="D9D9D9", fill_type="solid")
            else:
                celda.font = fuente_datos
                if opciones.get("centrar_texto", True):
                    celda.alignment = Alignment(horizontal="center")

    if opciones.get("inmovilizar_fila_1", True):
        ws.freeze_panes = "A2"

    if opciones.get("autocolumnas", True):
        for col in ws.columns:
            max_length = 0
            columna = col[0].column_letter
            for celda in col:
                try:
                    if celda.value:
                        max_length = max(max_length, len(str(celda.value)))
                except:
                    pass
            ajuste = min((max_length + 2), 50)
            ws.column_dimensions[columna].width = ajuste

    wb.save(ruta_excel)

def verificar_existencia_archivo(ruta):
    if os.path.exists(ruta):
        nuevo_nombre = messagebox.askquestion("Archivo existente", f"El archivo '{os.path.basename(ruta)}' ya existe.\n\n¿Desea sobrescribirlo? (No para elegir otro nombre)")
        if nuevo_nombre == "no":
            base, ext = os.path.splitext(ruta)
            ruta = base + "_nuevo" + ext
    return ruta

def verificar_tamanio(df):
    if len(df) > MAX_FILAS_EXCEL:
        opcion = messagebox.askquestion(
            "Archivo muy grande",
            f"El archivo tiene más de {MAX_FILAS_EXCEL:,} filas.\n\n¿Desea fragmentarlo en hojas dentro del mismo archivo? (No = fragmentar en archivos separados)")
        if opcion == "yes":
            return "hojas"
        else:
            return "archivos"
    return None

def convertir_basico(ruta_entrada, ruta_salida, quitar_x):
    from utilidades import registrar_evento
    nombre_base = obtener_nombre_archivo_sin_extension(ruta_entrada)
    if quitar_x and nombre_base.lower().startswith("x"):
        nombre_base = nombre_base[1:]
    ruta_salida = os.path.join(ruta_salida, f"{nombre_base}.xlsx")
    ruta_salida = verificar_existencia_archivo(ruta_salida)
    try:
        df = convertir_dbf_a_dataframe(ruta_entrada)
        fragmentar = verificar_tamanio(df)
        guardar_dataframe_como_excel(df, ruta_salida,
                                     fragmento_hojas=(fragmentar == "hojas"),
                                     fragmento_archivos=(fragmentar == "archivos"))
        registrar_evento(f"Dbf to Excel - 1 - EXITO")
    except Exception as e:
        registrar_evento(f"Dbf to Excel - 1 - FALLO - {e}")

def convertir_avanzado(ruta_entrada, ruta_salida, opciones):
    from utilidades import registrar_evento
    nombre_base = obtener_nombre_archivo_sin_extension(ruta_entrada)
    if opciones.get("quitar_x", False) and nombre_base.lower().startswith("x"):
        nombre_base = nombre_base[1:]
    ruta_salida = os.path.join(ruta_salida, f"{nombre_base}.xlsx")
    ruta_salida = verificar_existencia_archivo(ruta_salida)
    try:
        df = convertir_dbf_a_dataframe(ruta_entrada)
        fragmentar = verificar_tamanio(df)
        guardar_dataframe_como_excel(df, ruta_salida,
                                     fragmento_hojas=(fragmentar == "hojas"),
                                     fragmento_archivos=(fragmentar == "archivos"))
        aplicar_personalizacion_excel(ruta_salida, opciones)
        registrar_evento(f"Dbf to Excel Avanzado - 1 - EXITO")
    except Exception as e:
        registrar_evento(f"Dbf to Excel Avanzado - 1 - FALLO - {e}")

def convertir_total(lista_rutas, ruta_salida_carpeta, opciones):
    from utilidades import registrar_evento
    exitos = 0
    fallos = 0
    for ruta_entrada in lista_rutas:
        nombre_archivo = obtener_nombre_archivo_sin_extension(ruta_entrada)
        if opciones.get("quitar_x", False) and nombre_archivo.lower().startswith("x"):
            nombre_archivo = nombre_archivo[1:]
        ruta_salida = os.path.join(ruta_salida_carpeta, f"{nombre_archivo}.xlsx")
        ruta_salida = verificar_existencia_archivo(ruta_salida)
        try:
            df = convertir_dbf_a_dataframe(ruta_entrada)
            fragmentar = verificar_tamanio(df)
            guardar_dataframe_como_excel(df, ruta_salida,
                                         fragmento_hojas=(fragmentar == "hojas"),
                                         fragmento_archivos=(fragmentar == "archivos"))
            aplicar_personalizacion_excel(ruta_salida, opciones)
            exitos += 1
        except Exception as e:
            registrar_evento(f"Dbf to Excel Total - 1 - FALLO - {ruta_entrada}: {e}")
            fallos += 1
    if fallos:
        registrar_evento(f"Dbf to Excel Total - {len(lista_rutas)} - FALLO - {fallos} fallos")
    else:
        registrar_evento(f"Dbf to Excel Total - {len(lista_rutas)} - EXITO")
    
def guardar_dataframe_como_csv(df, ruta_salida):
    df.to_csv(ruta_salida, index=False)